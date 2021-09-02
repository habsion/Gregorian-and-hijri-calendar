import calendar
from hijri_converter import Gregorian
import openpyxl

def hijri_month_range(a):
    return (month_dict[int(str(Gregorian(y, m, 1).to_hijri()).split('-')[1])]+' - '+month_dict[int(str(Gregorian(y, m, a).to_hijri()).split('-')[1])]).center(20)


month_dict = {1: 'محرم', 2: 'صفر', 3: 'ربيع الأول', 4: 'ربيع الثاني', 5: 'جمادي الأولى', 6: 'جمادي الآخرة', 7: 'رجب'
, 8: 'شعبان', 9: 'رمضان', 10: 'شوال', 11: 'ذو القعدة', 12: 'ذو الحجة'}

y=int(input('Enter year: '))
wb=openpyxl.load_workbook('D:\\Python\\Calendar.xlsx')
for m in range(1,13):
    mon=calendar.month(y,m).split('\n')
    sheet=wb[str(m)]
    sheet.cell(1,4,value=mon[0].strip())
    sheet.cell(2,4,hijri_month_range(int(mon[-2].rstrip()[-2::])))
    mon[1]=mon[1].split()
    for i in range(1,8):
        sheet.cell(3,i,value=mon[1][i-1])
    mon[2]=mon[2].split()
    mon[2]=['']*(7-len(mon[2]))+mon[2]
    for i in range(1,8):
        if mon[2][i-1] != '':
            sheet.cell(4,i,mon[2][i-1]+' | '+str(Gregorian(y, m, int(mon[2][i-1])).to_hijri()).split('-')[2])
    for i in range(5,5+len(mon[3:-1])):
        mon[i-2]=mon[i-2].split()
        for j in range(1,len(mon[i-2])+1):
            sheet.cell(i,j,mon[i-2][j-1]+' | '+str(Gregorian(y, m, int(mon[i-2][j-1])).to_hijri()).split('-')[2])
wb.save('D:\\Python\\Calendar.xlsx')
